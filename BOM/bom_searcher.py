from __future__ import annotations

import argparse
import re
import sys
import tempfile
import threading
import traceback
from collections import defaultdict
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
from time import perf_counter
from typing import Any, Callable, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    import pythoncom
    from win32com.client import DispatchEx
except ImportError:
    pythoncom = None
    DispatchEx = None


SEARCHABLE_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm", ".xls"}
VERSION_TOKEN_PATTERN = re.compile(r"(?<!\d)(\d{6,8})(?!\d)")
IGNORED_SHEET_NAME_TOKENS = {
    "搜索结果",
    "汇总bom",
    "搜索明细",
    "未找到",
    "onlymanualseq",
    "onlysearchseq",
    "drawingdelta",
    "notfound",
}
REPORT_WORKBOOK_TOKENS = {
    "vssearchdiff",
    "searchdiff",
    "drawingdelta",
    "onlymanualseq",
    "onlysearchseq",
    "notfound",
    "compare",
    "对比",
    "差异",
    "比对",
}
HEADER_ALIASES = {
    "sequence": {"序号", "序", "项次", "項次"},
    "material_id": {"物料id", "料号", "料號", "物料编号", "物料編號", "物料编码", "物料編碼"},
    "drawing": {"图号", "圖號", "零件图号", "零件號", "图纸号", "图纸编号", "部件图号"},
    "name": {"名称", "品名", "零件名称", "部件名称"},
    "thickness": {"厚度", "厚", "板厚", "规格"},
    "material": {"材质", "材料"},
    "quantity": {"数量", "件数", "用量"},
    "total_quantity": {"总数量", "總數量", "总数", "合计数量", "总量"},
    "remark": {"备注", "備註", "说明", "說明"},
}
BOM_RESULT_HEADERS = ["序号", "物料ID", "图号", "名称", "厚度", "材质", "数量", "总数量", "备注"]
RESULT_HEADERS = [
    "输入工作表",
    "输入行号",
    "输入序号",
    "输入图号",
    "输入名称",
    "输入数量",
    "来源文件",
    "来源工作表",
    "父行号",
    "父序号",
    "父图号",
    "父名称",
    "子行号",
    "子序号",
    "子图号",
    "子名称",
    "厚度",
    "材质",
    "数量",
    "总数量",
    "备注",
    "结果状态",
]
NOT_FOUND_HEADERS = [
    "输入工作表",
    "输入行号",
    "输入序号",
    "输入图号",
    "输入名称",
    "输入数量",
    "状态",
]


@dataclass
class RowRecord:
    excel_row: int
    sequence: str
    material_id: str
    drawing: str
    drawing_base: str
    name: str
    thickness: str
    material: str
    quantity: str
    total_quantity: str
    remark: str
    raw_values: tuple[object, ...] = field(default_factory=tuple)

    def has_payload(self) -> bool:
        return any(
            [
                self.drawing,
                self.name,
                self.material_id,
                self.thickness,
                self.material,
                self.quantity,
                self.total_quantity,
                self.remark,
            ]
        )


@dataclass
class QueryItem:
    source_sheet: str
    source_row: int
    sequence: str
    drawing: str
    drawing_base: str
    name: str
    quantity: str


@dataclass
class SheetTable:
    workbook_path: Path
    sheet_name: str
    header_row: int
    header_values: tuple[object, ...]
    columns: dict[str, int]
    rows: list[RowRecord]


@dataclass
class SearchResult:
    query: QueryItem
    table: SheetTable
    parent: RowRecord
    child: RowRecord


@dataclass
class SearchGroup:
    query: QueryItem
    table: SheetTable
    parent: RowRecord
    children: list[RowRecord]


@dataclass
class SearchSummary:
    total_queries: int
    total_results: int
    not_found_count: int
    output_path: Path
    scanned_files: int
    skipped_files: list[str]


@dataclass
class AggregatedBOMItem:
    source_sheet: str
    source_row: int
    sequence: str
    material_id: str
    drawing: str
    name: str
    thickness: str
    material: str
    remark: str
    quantity_values: set[str] = field(default_factory=set)
    total_quantity_decimal: Decimal = Decimal("0")
    has_numeric_total: bool = False
    total_quantity_text: str = ""

    def add_occurrence(self, quantity: str, total_quantity: str) -> None:
        quantity = normalize_text(quantity)
        if quantity:
            self.quantity_values.add(quantity)

        total_quantity = normalize_text(total_quantity)
        total_decimal = parse_decimal(total_quantity)
        if total_decimal is not None:
            self.total_quantity_decimal += total_decimal
            self.has_numeric_total = True
            return

        if not total_quantity or self.has_numeric_total:
            return

        if not self.total_quantity_text:
            self.total_quantity_text = total_quantity
        elif self.total_quantity_text != total_quantity:
            self.total_quantity_text = ""

    def display_quantity(self) -> str:
        if len(self.quantity_values) == 1:
            return next(iter(self.quantity_values))
        return ""

    def display_total_quantity(self) -> str:
        if self.has_numeric_total:
            return format_decimal(self.total_quantity_decimal)
        return self.total_quantity_text


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value}".rstrip("0").rstrip(".")
    if isinstance(value, int):
        return str(value)
    text = str(value)
    translation = str.maketrans(
        {
            "（": "(",
            "）": ")",
            "－": "-",
            "—": "-",
            "　": " ",
            "\n": " ",
            "\r": " ",
            "\t": " ",
        }
    )
    return text.translate(translation).strip()


def normalize_header(value: object) -> str:
    return re.sub(r"\s+", "", normalize_text(value)).lower()


def normalize_name_token(value: object) -> str:
    return re.sub(r"[\s_-]+", "", normalize_text(value)).lower()


def trim_trailing_empty(values: Iterable[object]) -> tuple[object, ...]:
    trimmed = list(values)
    while trimmed and not normalize_text(trimmed[-1]):
        trimmed.pop()

    # Some source BOM files contain a single stray value in the far-right tail
    # (for example XFD) after a very long gap of empty cells. Keeping that tail
    # would expand the exported raw sheet to thousands of columns and make
    # formatting appear stuck. Drop isolated sparse tails while preserving
    # normal nearby extra columns.
    while True:
        nonempty_positions = [index for index, value in enumerate(trimmed) if normalize_text(value)]
        if len(nonempty_positions) < 2:
            break
        if nonempty_positions[-1] - nonempty_positions[-2] <= 32:
            break
        trimmed = trimmed[: nonempty_positions[-2] + 1]

    return tuple(trimmed)


def value_at(values: tuple[object, ...], index: int | None) -> object:
    if index is None or index < 0 or index >= len(values):
        return None
    return values[index]


def normalize_sequence(value: object) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    text = re.sub(r"\s+", "", text)
    return text.rstrip(".")


def extract_drawing_base(value: object) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    compact = re.sub(r"\s+", "", text)
    match = re.match(r"([A-Za-z0-9-]+)", compact)
    if match:
        return match.group(1).upper()
    match = re.search(r"([A-Za-z0-9-]{4,})", compact)
    if match:
        return match.group(1).upper()
    return compact.upper()


def parse_decimal(value: object) -> Decimal | None:
    text = normalize_text(value)
    if not text:
        return None
    text = text.replace(",", "")
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def format_decimal(value: Decimal) -> str:
    text = format(value.normalize(), "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text


def extract_latest_version(value: object) -> int:
    text = normalize_text(value)
    best = 0
    for token in VERSION_TOKEN_PATTERN.findall(text):
        if len(token) == 6:
            normalized = int(f"20{token}")
        else:
            normalized = int(token)
        best = max(best, normalized)
    return best


def row_is_empty(values: Iterable[object]) -> bool:
    return not any(normalize_text(value) for value in values)


def detect_header(rows: list[tuple[object, ...]], max_scan_rows: int = 30) -> tuple[int, dict[str, int]] | None:
    best_score = -1
    best_match: tuple[int, dict[str, int]] | None = None

    for row_index, row in enumerate(rows[:max_scan_rows], start=1):
        columns: dict[str, int] = {}
        score = 0
        for column_index, cell_value in enumerate(row):
            header = normalize_header(cell_value)
            if not header:
                continue
            for field, aliases in HEADER_ALIASES.items():
                if header in aliases and field not in columns:
                    columns[field] = column_index
                    score += 3 if field == "drawing" else 1
                    break
        if "drawing" in columns and score > best_score:
            best_score = score
            best_match = (row_index, columns)

    return best_match


def make_row_record(values: tuple[object, ...], excel_row: int, columns: dict[str, int]) -> RowRecord:
    def value_for(field: str) -> str:
        column_index = columns.get(field)
        if column_index is None or column_index >= len(values):
            return ""
        return normalize_text(values[column_index])

    sequence = value_for("sequence")
    drawing = value_for("drawing")
    return RowRecord(
        excel_row=excel_row,
        sequence=normalize_sequence(sequence),
        material_id=value_for("material_id"),
        drawing=drawing,
        drawing_base=extract_drawing_base(drawing),
        name=value_for("name"),
        thickness=value_for("thickness"),
        material=value_for("material"),
        quantity=value_for("quantity"),
        total_quantity=value_for("total_quantity"),
        remark=value_for("remark"),
        raw_values=trim_trailing_empty(values),
    )


def parse_sheet(path: Path, sheet_name: str, rows: list[tuple[object, ...]]) -> SheetTable | None:
    if normalize_name_token(sheet_name) in IGNORED_SHEET_NAME_TOKENS:
        return None

    header_info = detect_header(rows)
    if not header_info:
        return None

    header_row, columns = header_info
    records: list[RowRecord] = []
    for offset, values in enumerate(rows[header_row:], start=header_row + 1):
        record = make_row_record(values, excel_row=offset, columns=columns)
        if record.has_payload() or record.sequence:
            records.append(record)

    return SheetTable(
        workbook_path=path,
        sheet_name=sheet_name,
        header_row=header_row,
        header_values=trim_trailing_empty(rows[header_row - 1]),
        columns=columns,
        rows=records,
    )


def collect_queries(input_file: Path) -> list[QueryItem]:
    workbook = load_workbook(input_file, data_only=True, read_only=True)
    queries: list[QueryItem] = []

    try:
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            table = parse_sheet(input_file, sheet.title, rows)
            if table is None:
                continue
            for row in table.rows:
                if not row.drawing_base:
                    continue
                queries.append(
                    QueryItem(
                        source_sheet=sheet.title,
                        source_row=row.excel_row,
                        sequence=row.sequence,
                        drawing=row.drawing,
                        drawing_base=row.drawing_base,
                        name=row.name,
                        quantity=row.quantity,
                    )
                )
    finally:
        workbook.close()

    return queries


def maybe_convert_xls(path: Path, temp_dir: Path) -> Path:
    if path.suffix.lower() != ".xls":
        return path
    if DispatchEx is None:
        raise RuntimeError("当前环境缺少 Excel COM 支持，无法读取 .xls 文件。")

    output_path = temp_dir / f"{path.stem}_converted.xlsx"
    if pythoncom is not None:
        pythoncom.CoInitialize()
    excel = None
    workbook = None
    try:
        excel = DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(str(path))
        workbook.SaveAs(str(output_path), FileFormat=51)
        return output_path
    finally:
        if workbook is not None:
            workbook.Close(False)
        if excel is not None:
            excel.Quit()
        if pythoncom is not None:
            pythoncom.CoUninitialize()


def load_tables_from_workbook(path: Path, temp_dir: Path) -> list[SheetTable]:
    actual_path = maybe_convert_xls(path, temp_dir)
    workbook = load_workbook(actual_path, data_only=True, read_only=True)
    tables: list[SheetTable] = []

    try:
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            table = parse_sheet(path, sheet.title, rows)
            if table is not None:
                tables.append(table)
    finally:
        workbook.close()

    return tables


def collect_workbook_paths(sources: Path | list[Path], excludes: set[Path]) -> list[Path]:
    files: list[Path] = []
    source_list = [sources] if isinstance(sources, Path) else sources

    for src in source_list:
        paths_to_check = src.rglob("*") if src.is_dir() else [src]
        for path in paths_to_check:
            if not path.is_file():
                continue
            if path.name.startswith("~$"):
                continue
            lower_stem = path.stem.lower()
            if "搜索结果" in path.stem or lower_stem.startswith("_") or lower_stem.startswith("result") or lower_stem.startswith("output"):
                continue
            normalized_stem = normalize_name_token(path.stem)
            if any(token in normalized_stem for token in REPORT_WORKBOOK_TOKENS):
                continue
            if path.suffix.lower() not in SEARCHABLE_EXTENSIONS:
                continue
            try:
                if path.resolve() in excludes:
                    continue
            except Exception:
                pass
            files.append(path)

    unique_files = []
    seen = set()
    for f in files:
        try:
            res = f.resolve()
        except Exception:
            res = f
        if res not in seen:
            seen.add(res)
            unique_files.append(f)

    return sorted(unique_files)


def is_descendant_sequence(child: str, parent: str) -> bool:
    if not child or not parent:
        return False
    # If child strictly starts with "parent.", it's a descendant.
    # We also handle the case where parent is "3.1" and child is "1.1" 
    # but the context means it's a sub-child. However, contextually,
    # the function extract_children is used to find all rows belonging
    # to the current group until the next sibling/uncle.
    return child.startswith(parent + ".")

def is_natural_sequence(sequence: str) -> bool:
    return bool(sequence) and bool(re.fullmatch(r"\d+", sequence))


def has_component_marker(row: RowRecord) -> bool:
    markers = {normalize_text(row.thickness), normalize_text(row.material)}
    return "组件" in markers


def has_subset_marker(row: RowRecord) -> bool:
    subset_fields = (
        row.name,
        row.thickness,
        row.material,
        row.remark,
    )
    return any("子集" in normalize_text(value) for value in subset_fields if value)


def prefer_explicit_sequence_candidates(
    candidates: list[tuple[SheetTable, int]],
) -> list[tuple[SheetTable, int]]:
    explicit_candidates = [
        (table, row_index)
        for table, row_index in candidates
        if table.rows[row_index].sequence
    ]
    return explicit_candidates or candidates


def prefer_natural_sequence_candidates(
    candidates: list[tuple[SheetTable, int]],
) -> list[tuple[SheetTable, int]]:
    natural_candidates = [
        (table, row_index)
        for table, row_index in candidates
        if is_natural_sequence(table.rows[row_index].sequence)
    ]
    return natural_candidates or candidates


def quantity_match_priority(group: SearchGroup) -> tuple[int, Decimal]:
    query_quantity = parse_decimal(group.query.quantity)
    parent_total_quantity = parse_decimal(group.parent.total_quantity)
    if query_quantity is None or parent_total_quantity is None:
        return (0, Decimal("-Infinity"))
    difference = abs(query_quantity - parent_total_quantity)
    return (int(difference == 0), -difference)

def candidate_priority(group: SearchGroup) -> tuple[int, int, int, Decimal, int, int, int]:
    query_name = normalize_header(group.query.name)
    parent_name = normalize_header(group.parent.name)
    name_match = int(bool(query_name) and query_name == parent_name)
    file_version = extract_latest_version(group.table.workbook_path.stem)
    quantity_exact_match, quantity_distance = quantity_match_priority(group)
    remark_version = extract_latest_version(group.parent.remark)
    has_children = int(bool(group.children))
    child_count = len(group.children)
    return (
        name_match,
        file_version,
        quantity_exact_match,
        quantity_distance,
        remark_version,
        has_children,
        child_count,
    )


def describe_candidate(group: SearchGroup) -> str:
    return (
        f"{group.table.workbook_path.name}"
        f"/{group.table.sheet_name}"
        f" 行{group.parent.excel_row}"
    )


def select_best_groups(
    query: QueryItem,
    candidate_groups: list[SearchGroup],
    log: Callable[[str], None] | None = None,
) -> list[SearchGroup]:
    if len(candidate_groups) <= 1:
        return candidate_groups

    best_group = max(candidate_groups, key=candidate_priority)
    if log:
        ignored = [describe_candidate(group) for group in candidate_groups if group is not best_group]
        log(
            f"图号 {query.drawing} 命中 {len(candidate_groups)} 个来源，"
            f"已保留: {describe_candidate(best_group)}；"
            f"忽略: {'；'.join(ignored)}"
        )
    return [best_group]


def extract_children(rows: list[RowRecord], parent_index: int) -> list[RowRecord]:
    parent = rows[parent_index]
    children: list[RowRecord] = []

    # Only rows explicitly marked as components are expanded. All other rows are
    # treated as leaf parts even if they happen to be followed by detail rows.
    if not has_component_marker(parent):
        return children

    # If the parent sequence is a natural number (e.g., "1", "2") or empty,
    # we collect everything until the next natural number sequence.
    if is_natural_sequence(parent.sequence) or not parent.sequence:
        for row in rows[parent_index + 1 :]:
            if is_natural_sequence(row.sequence):
                break
            if row.has_payload() or row.sequence:
                children.append(row)
        return children

    # For parent sequences like "18.2", only rows nested under that exact
    # sequence belong to the same component. Once another explicit sequence
    # appears, the component boundary has ended.
    parent_sequence = parent.sequence
    blank_streak = 0

    for row in rows[parent_index + 1 :]:
        if not row.has_payload() and not row.sequence:
            blank_streak += 1
            if blank_streak >= 2 and children:
                break
            continue

        blank_streak = 0
        current_sequence = row.sequence

        if current_sequence:
            if is_descendant_sequence(current_sequence, parent_sequence):
                children.append(row)
                continue

            break

        if row.has_payload():
            children.append(row)

    return children


def build_index(
    workbook_paths: list[Path],
    log: Callable[[str], None] | None = None,
) -> tuple[dict[str, list[tuple[SheetTable, int]]], list[str]]:
    index: dict[str, list[tuple[SheetTable, int]]] = defaultdict(list)
    skipped: list[str] = []

    with tempfile.TemporaryDirectory(prefix="bom_searcher_") as temp_root:
        temp_dir = Path(temp_root)
        for path in workbook_paths:
            try:
                tables = load_tables_from_workbook(path, temp_dir)
            except Exception as exc:
                skipped.append(f"{path.name}: {exc}")
                if log:
                    log(f"跳过文件: {path.name}，原因: {exc}")
                continue

            matched_rows = 0
            for table in tables:
                for row_index, row in enumerate(table.rows):
                    if not row.drawing_base:
                        continue
                    index[row.drawing_base].append((table, row_index))
                    matched_rows += 1

            if log:
                log(f"已扫描: {path.name}，索引到 {matched_rows} 条候选图号")

    return index, skipped


def search_queries(
    queries: list[QueryItem],
    index: dict[str, list[tuple[SheetTable, int]]],
    log: Callable[[str], None] | None = None,
) -> tuple[list[SearchGroup], list[tuple[QueryItem, str]]]:
    groups: list[SearchGroup] = []
    not_found: list[tuple[QueryItem, str]] = []

    for query in queries:
        candidates = index.get(query.drawing_base, [])
        candidates = prefer_explicit_sequence_candidates(candidates)
        candidates = prefer_natural_sequence_candidates(candidates)
        if not candidates:
            not_found.append((query, "未找到匹配图号"))
            continue

        candidate_groups: list[SearchGroup] = []
        for table, parent_index in candidates:
            parent = table.rows[parent_index]
            children = extract_children(table.rows, parent_index)
            candidate_groups.append(
                SearchGroup(
                    query=query,
                    table=table,
                    parent=parent,
                    children=children,
                )
            )
        candidate_groups = select_best_groups(query, candidate_groups, log=log)
        groups.extend(candidate_groups)

        if not candidate_groups:
            not_found.append((query, "找到父图号，但未提取到子项，请检查序号列"))

    return groups, not_found


def autosize_columns(sheet) -> None:
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            display_length = sum(2 if ord(char) > 127 else 1 for char in value)
            max_length = max(max_length, display_length)
        sheet.column_dimensions[column_letter].width = min(max_length + 2, 40)


def style_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    autosize_columns(sheet)


def configure_a4_landscape(sheet) -> None:
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0


def get_display_sequence(group: SearchGroup, row: RowRecord, *, is_parent: bool) -> str:
    query_seq = group.query.sequence
    bom_parent_seq = group.parent.sequence

    if is_parent:
        return query_seq or bom_parent_seq

    child_seq = row.sequence
    
    # If the child has no sequence, generate one automatically
    # like "6.1", "6.2", etc., based on the parent's query sequence
    if not child_seq:
        # Find index of this child to determine its automatic suffix
        try:
            child_idx = group.children.index(row) + 1
            base_seq = query_seq or bom_parent_seq or ""
            if base_seq:
                return f"{base_seq}.{child_idx}"
            return str(child_idx)
        except ValueError:
            return ""

    if query_seq and bom_parent_seq:
        if child_seq.startswith(bom_parent_seq + "."):
            return query_seq + child_seq[len(bom_parent_seq):]
        
        # Handle case where parent is "3.1" but child is "1.1"
        # We want to replace the top-level prefix of child_seq
        # with the query_seq. E.g. query_seq="5.1", child_seq="1.1" -> "5.1.1"
        # We assume if child_seq has dots, it's a sub-sequence.
        if "." in child_seq:
            # e.g., child is "1.1", query is "3.1".
            # It should become "3.1.1"
            parts = child_seq.split(".", 1)
            if len(parts) == 2:
                return f"{query_seq}.{parts[1]}"
        
        # If no dots, just append it like "3.1.1"
        return f"{query_seq}.{child_seq}"

    return child_seq


def get_display_total_quantity(group: SearchGroup, row: RowRecord, *, is_parent: bool) -> str:
    query_total = normalize_text(group.query.quantity)
    if is_parent:
        return query_total or row.total_quantity

    query_decimal = parse_decimal(query_total)
    child_total_decimal = parse_decimal(row.total_quantity)
    parent_total_decimal = parse_decimal(group.parent.total_quantity)

    if (
        child_total_decimal is not None
        and query_decimal is not None
        and parent_total_decimal not in (None, Decimal("0"))
    ):
        return format_decimal((child_total_decimal * query_decimal) / parent_total_decimal)

    child_quantity_decimal = parse_decimal(row.quantity)
    if child_quantity_decimal is not None and query_decimal is not None:
        return format_decimal(child_quantity_decimal * query_decimal)

    return query_total or row.total_quantity


def get_display_quantity(group: SearchGroup, row: RowRecord, *, is_parent: bool) -> str:
    if is_parent:
        return "1"
    return row.quantity


def make_display_row(group: SearchGroup, row: RowRecord, *, is_parent: bool) -> list[str]:
    return [
        get_display_sequence(group, row, is_parent=is_parent),
        row.material_id,
        row.drawing,
        row.name,
        row.thickness,
        row.material,
        get_display_quantity(group, row, is_parent=is_parent),
        get_display_total_quantity(group, row, is_parent=is_parent),
        row.remark,
    ]


def get_summary_total_quantity(group: SearchGroup, row: RowRecord, *, is_parent: bool) -> str:
    if is_parent:
        # The summary sheet is task-order-driven: only the task quantity is
        # retained in the total-quantity column for the matched parent row.
        return normalize_text(group.query.quantity)
    return ""


def make_summary_row(group: SearchGroup, row: RowRecord, *, is_parent: bool) -> list[str]:
    return [
        get_display_sequence(group, row, is_parent=is_parent),
        row.material_id,
        row.drawing,
        row.name,
        row.thickness,
        row.material,
        get_display_quantity(group, row, is_parent=is_parent),
        get_summary_total_quantity(group, row, is_parent=is_parent),
        row.remark,
    ]


def get_raw_header(groups: list[SearchGroup]) -> tuple[object, ...]:
    best_header: tuple[object, ...] = ()
    for group in groups:
        header = group.table.header_values
        if len(header) > len(best_header):
            best_header = header
    return best_header or tuple(BOM_RESULT_HEADERS)


def get_raw_row_width(groups: list[SearchGroup]) -> int:
    width = len(get_raw_header(groups))
    for group in groups:
        rows = [group.parent, *group.children]
        for row in rows:
            width = max(width, len(row.raw_values))
    return width


def pad_row(values: tuple[object, ...], width: int) -> list[object]:
    return list(values) + [None] * max(0, width - len(values))


def is_aggregatable_row(row: RowRecord) -> bool:
    return bool(row.material_id or row.drawing_base)


def should_aggregate_summary_row(row: RowRecord, *, allow_component: bool) -> bool:
    if not is_aggregatable_row(row):
        return False
    if has_subset_marker(row):
        return False
    if not allow_component and has_component_marker(row):
        return False
    return True


def get_summary_sequence(group: SearchGroup) -> str:
    return normalize_sequence(group.query.sequence) or normalize_sequence(group.parent.sequence)


def should_allow_parent_summary_fallback(group: SearchGroup) -> bool:
    if not has_component_marker(group.parent):
        return True
    if has_subset_marker(group.parent):
        return False
    parent_sequence = normalize_sequence(group.parent.sequence)
    query_sequence = normalize_sequence(group.query.sequence)
    if "." in parent_sequence or "." in query_sequence:
        return False
    return True


def aggregate_groups(groups: list[SearchGroup]) -> list[AggregatedBOMItem]:
    aggregated: dict[
        tuple[str, int, str, str, str, str, str, str],
        AggregatedBOMItem,
    ] = {}

    for group in groups:
        use_parent_fallback = not group.children
        summary_sequence = get_summary_sequence(group)
        if group.children:
            rows = [
                row
                for row in group.children
                if should_aggregate_summary_row(row, allow_component=False)
            ]
        else:
            rows = (
                [group.parent]
                if should_aggregate_summary_row(
                    group.parent,
                    allow_component=should_allow_parent_summary_fallback(group),
                )
                else []
            )
        for row in rows:
            key = (
                group.query.source_sheet,
                group.query.source_row,
                summary_sequence,
                row.material_id,
                row.drawing_base,
                row.name,
                row.thickness,
                row.material,
                row.remark,
            )
            item = aggregated.get(key)
            if item is None:
                item = AggregatedBOMItem(
                    source_sheet=group.query.source_sheet,
                    source_row=group.query.source_row,
                    sequence=summary_sequence,
                    material_id=row.material_id,
                    drawing=row.drawing,
                    name=row.name,
                    thickness=row.thickness,
                    material=row.material,
                    remark=row.remark,
                )
                aggregated[key] = item
            quantity = get_display_quantity(group, row, is_parent=use_parent_fallback)
            total_quantity = get_display_total_quantity(group, row, is_parent=use_parent_fallback)
            item.add_occurrence(quantity, total_quantity)

    return list(aggregated.values())


def get_sequence_cell_value(sequence: str) -> object:
    sequence = normalize_sequence(sequence)
    if is_natural_sequence(sequence):
        try:
            return int(sequence)
        except ValueError:
            return sequence
    return sequence


def should_include_summary_child_row(row: RowRecord) -> bool:
    return row.has_payload() or bool(row.sequence)


def write_bom_like_sheet(sheet, groups: list[SearchGroup]) -> int:
    sheet.append(BOM_RESULT_HEADERS)
    parent_rows: list[int] = []
    result_row_count = 0

    for group in groups:
        parent_rows.append(sheet.max_row + 1)
        sheet.append(make_display_row(group, group.parent, is_parent=True))
        result_row_count += 1
        for child in group.children:
            sheet.append(make_display_row(group, child, is_parent=False))
            result_row_count += 1

    style_sheet(sheet)
    parent_font = Font(color="C00000", bold=True)
    for row_index in parent_rows:
        for cell in sheet[row_index]:
            cell.font = parent_font
    return result_row_count


def write_raw_bom_sheet(sheet, groups: list[SearchGroup]) -> int:
    standard_fields = (
        "sequence",
        "material_id",
        "drawing",
        "name",
        "thickness",
        "material",
        "quantity",
        "total_quantity",
        "remark",
    )
    extra_headers: list[tuple[str, str]] = []
    seen_extra_keys: set[str] = set()

    for group in groups:
        table = group.table
        standard_indexes = set(table.columns.values())
        for index, header_value in enumerate(table.header_values):
            if index in standard_indexes:
                continue
            display_header = normalize_text(header_value)
            if not display_header:
                continue
            key = normalize_header(display_header)
            if not key or key in seen_extra_keys:
                continue
            seen_extra_keys.add(key)
            extra_headers.append((key, display_header))

    sheet.append([*BOM_RESULT_HEADERS, *[label for _, label in extra_headers]])

    parent_rows: list[int] = []
    result_row_count = 0

    for group in groups:
        table = group.table
        standard_indexes = set(table.columns.values())

        def build_row(row: RowRecord) -> list[object]:
            result = [
                value_at(row.raw_values, table.columns.get(field))
                for field in standard_fields
            ]
            extra_values_by_key: dict[str, object] = {}
            for index, header_value in enumerate(table.header_values):
                if index in standard_indexes:
                    continue
                display_header = normalize_text(header_value)
                if not display_header:
                    continue
                key = normalize_header(display_header)
                if key and key not in extra_values_by_key:
                    extra_values_by_key[key] = value_at(row.raw_values, index)
            for key, _ in extra_headers:
                result.append(extra_values_by_key.get(key))
            return result

        parent_rows.append(sheet.max_row + 1)
        sheet.append(build_row(group.parent))
        result_row_count += 1
        for child in group.children:
            sheet.append(build_row(child))
            result_row_count += 1

    style_sheet(sheet)
    parent_font = Font(color="C00000", bold=True)
    for row_index in parent_rows:
        for cell in sheet[row_index]:
            cell.font = parent_font
    return result_row_count


def write_aggregated_sheet(sheet, groups: list[SearchGroup]) -> int:
    sheet.append(BOM_RESULT_HEADERS)
    parent_rows: list[int] = []
    result_row_count = 0

    for group in groups:
        parent_rows.append(sheet.max_row + 1)
        sheet.append(make_summary_row(group, group.parent, is_parent=True))
        result_row_count += 1
        for child in group.children:
            if not should_include_summary_child_row(child):
                continue
            sheet.append(make_summary_row(group, child, is_parent=False))
            result_row_count += 1

    style_sheet(sheet)
    parent_font = Font(color="C00000", bold=True)
    for row_index in parent_rows:
        for cell in sheet[row_index]:
            cell.font = parent_font
    configure_a4_landscape(sheet)
    return result_row_count


def write_detail_sheet(sheet, groups: list[SearchGroup]) -> int:
    sheet.append(RESULT_HEADERS)
    detail_count = 0

    for group in groups:
        if not group.children:
            detail_count += 1
            sheet.append(
                [
                    group.query.source_sheet,
                    group.query.source_row,
                    group.query.sequence,
                    group.query.drawing,
                    group.query.name,
                    group.query.quantity,
                    group.table.workbook_path.name,
                    group.table.sheet_name,
                    group.parent.excel_row,
                    group.parent.sequence,
                    group.parent.drawing,
                    group.parent.name,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    "仅父项，无子项",
                ]
            )
            continue

        for child in group.children:
            detail_count += 1
            sheet.append(
                [
                    group.query.source_sheet,
                    group.query.source_row,
                    group.query.sequence,
                    group.query.drawing,
                    group.query.name,
                    group.query.quantity,
                    group.table.workbook_path.name,
                    group.table.sheet_name,
                    group.parent.excel_row,
                    group.parent.sequence,
                    group.parent.drawing,
                    group.parent.name,
                    child.excel_row,
                    child.sequence,
                    child.drawing,
                    child.name,
                    child.thickness,
                    child.material,
                    child.quantity,
                    child.total_quantity,
                    child.remark,
                    "已提取子项",
                ]
            )

    style_sheet(sheet)
    return detail_count


def write_not_found_sheet(sheet, not_found: list[tuple[QueryItem, str]]) -> None:
    sheet.append(NOT_FOUND_HEADERS)
    for query, reason in not_found:
        sheet.append(
            [
                query.source_sheet,
                query.source_row,
                query.sequence,
                query.drawing,
                query.name,
                query.quantity,
                reason,
            ]
        )
    style_sheet(sheet)


def write_output(
    output_path: Path,
    groups: list[SearchGroup],
    not_found: list[tuple[QueryItem, str]],
) -> int:
    workbook = Workbook()
    try:
        aggregated_sheet = workbook.active
        aggregated_sheet.title = "汇总BOM"
        write_aggregated_sheet(aggregated_sheet, groups)

        result_sheet = workbook.create_sheet("搜索结果")
        result_row_count = write_raw_bom_sheet(result_sheet, groups)

        detail_sheet = workbook.create_sheet("搜索明细")
        write_detail_sheet(detail_sheet, groups)

        not_found_sheet = workbook.create_sheet("未找到")
        write_not_found_sheet(not_found_sheet, not_found)

        workbook.active = 0
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        return result_row_count
    finally:
        workbook.close()


def run_search(
    input_file: Path,
    bom_sources: Path | list[Path],
    output_file: Path,
    log: Callable[[str], None] | None = None,
) -> SearchSummary:
    overall_start = perf_counter()
    input_file = input_file.resolve()
    output_file = output_file.resolve()

    if not input_file.exists():
        raise FileNotFoundError(f"输入文件不存在: {input_file}")
        
    source_list = [bom_sources] if isinstance(bom_sources, Path) else bom_sources
    for src in source_list:
        if not src.exists():
            raise FileNotFoundError(f"BOM 路径不存在: {src}")

    if log:
        log("正在读取任务单...")
    queries = collect_queries(input_file)
    query_end = perf_counter()
    if not queries:
        raise RuntimeError("任务单里没有识别到“图号”列或有效图号。")
    if log:
        log(f"任务单图号读取完成，共 {len(queries)} 条")

    workbook_paths = collect_workbook_paths(bom_sources, excludes={input_file, output_file})
    if not workbook_paths:
        raise RuntimeError("提供的 BOM 路径下没有可搜索的 Excel 文件。")
    if log:
        log(f"待扫描 BOM 文件数: {len(workbook_paths)}")

    index, skipped_files = build_index(workbook_paths, log=log)
    index_end = perf_counter()
    if log:
        log(f"索引完成，可匹配图号种类: {len(index)}")

    groups, not_found = search_queries(queries, index, log=log)
    search_end = perf_counter()
    result_row_count = sum(1 + len(group.children) for group in groups)
    if log:
        log(f"搜索完成，共找到 {result_row_count} 行结果，未命中 {len(not_found)} 条")
        log("正在保存结果文件...")

    write_output(output_file, groups, not_found)
    save_end = perf_counter()
    if log:
        log(f"结果已导出: {output_file}")
        log(
            "阶段耗时:"
            f" 读取任务单 {query_end - overall_start:.2f}s,"
            f" 建立索引 {index_end - query_end:.2f}s,"
            f" 检索 {search_end - index_end:.2f}s,"
            f" 保存结果 {save_end - search_end:.2f}s"
        )

    return SearchSummary(
        total_queries=len(queries),
        total_results=result_row_count,
        not_found_count=len(not_found),
        output_path=output_file,
        scanned_files=len(workbook_paths),
        skipped_files=skipped_files,
    )


class BOMSearcherApp:
    def __init__(self, root: Any, tk_module: Any, ttk_module: Any, filedialog_module: Any, messagebox_module: Any) -> None:
        self.root = root
        self.tk = tk_module
        self.ttk = ttk_module
        self.filedialog = filedialog_module
        self.messagebox = messagebox_module
        self.root.title("BOM 搜索器")
        self.root.geometry("900x620")

        self.input_var = self.tk.StringVar()
        self.bom_var = self.tk.StringVar()
        self.output_var = self.tk.StringVar()

        self._build_ui()

    def _build_ui(self) -> None:
        frame = self.ttk.Frame(self.root, padding=16)
        frame.pack(fill="both", expand=True)

        title = self.ttk.Label(frame, text="BOM 搜索器", font=("Microsoft YaHei UI", 16, "bold"))
        title.pack(anchor="w")

        desc = self.ttk.Label(
            frame,
            text="读取任务单的图号，再到指定的 BOM Excel 文件中抓取对应子件明细。",
        )
        desc.pack(anchor="w", pady=(4, 16))

        self._path_row(frame, "任务单", self.input_var, self.choose_input_file)
        self._path_row(frame, "BOM 文件", self.bom_var, self.choose_bom_files)
        self._path_row(frame, "输出结果", self.output_var, self.choose_output_file)

        button_bar = self.ttk.Frame(frame)
        button_bar.pack(fill="x", pady=(8, 12))

        self.run_button = self.ttk.Button(button_bar, text="开始搜索", command=self.start_search)
        self.run_button.pack(side="left")

        self.ttk.Button(button_bar, text="清空日志", command=self.clear_log).pack(side="left", padx=(8, 0))

        log_label = self.ttk.Label(frame, text="运行日志")
        log_label.pack(anchor="w", pady=(8, 6))

        self.log_text = self.tk.Text(frame, height=24, wrap="word", font=("Consolas", 10))
        self.log_text.pack(fill="both", expand=True)

    def _path_row(
        self,
        parent,
        label_text: str,
        variable: Any,
        command: Callable[[], None],
    ) -> None:
        row = self.ttk.Frame(parent)
        row.pack(fill="x", pady=6)
        self.ttk.Label(row, text=label_text, width=10).pack(side="left")
        self.ttk.Entry(row, textvariable=variable).pack(side="left", fill="x", expand=True, padx=(0, 8))
        self.ttk.Button(row, text="选择", command=command).pack(side="left")

    def choose_input_file(self) -> None:
        path = self.filedialog.askopenfilename(
            title="选择任务单",
            filetypes=[("Excel 文件", "*.xlsx *.xlsm *.xltx *.xltm *.xls")],
        )
        if not path:
            return
        self.input_var.set(path)
        if not self.output_var.get():
            suggested = Path(path).with_name(f"{Path(path).stem}_搜索结果.xlsx")
            self.output_var.set(str(suggested))

    def choose_bom_files(self) -> None:
        paths = self.filedialog.askopenfilenames(
            title="选择 BOM 文件 (可多选)",
            filetypes=[("Excel 文件", "*.xlsx *.xlsm *.xltx *.xltm *.xls")],
        )
        if paths:
            self.bom_var.set(";".join(paths))

    def choose_output_file(self) -> None:
        path = self.filedialog.askdirectory(title="选择输出结果保存的文件夹")
        if path:
            input_path = self.input_var.get().strip()
            if input_path:
                suggested_name = f"{Path(input_path).stem}_搜索结果.xlsx"
            else:
                suggested_name = "搜索结果.xlsx"
            self.output_var.set(str(Path(path) / suggested_name))

    def clear_log(self) -> None:
        self.log_text.delete("1.0", "end")

    def log(self, message: str) -> None:
        self.log_text.insert("end", message + "\n")
        self.log_text.see("end")
        self.root.update_idletasks()

    def start_search(self) -> None:
        input_path = self.input_var.get().strip()
        bom_paths_str = self.bom_var.get().strip()
        output_path = self.output_var.get().strip()

        if not input_path or not bom_paths_str or not output_path:
            self.messagebox.showwarning("提示", "请先选择任务单、BOM 文件和输出结果路径。")
            return

        bom_sources = [Path(p.strip()) for p in bom_paths_str.split(";") if p.strip()]

        self.run_button.config(state="disabled")
        self.log("开始执行搜索...")

        thread = threading.Thread(
            target=self._search_worker,
            args=(Path(input_path), bom_sources, Path(output_path)),
            daemon=True,
        )
        thread.start()

    def _search_worker(self, input_file: Path, bom_sources: list[Path], output_file: Path) -> None:
        try:
            summary = run_search(
                input_file=input_file,
                bom_sources=bom_sources,
                output_file=output_file,
                log=lambda msg: self.root.after(0, self.log, msg),
            )
            self.root.after(0, self._on_search_success, summary)
        except Exception as exc:
            detail = traceback.format_exc()
            self.root.after(0, self._on_search_failed, exc, detail)

    def _on_search_success(self, summary: SearchSummary) -> None:
        self.run_button.config(state="normal")
        if summary.skipped_files:
            self.log("以下文件未处理成功：")
            for item in summary.skipped_files:
                self.log(f"  - {item}")
        self.messagebox.showinfo(
            "完成",
            (
                f"搜索完成。\n"
                f"任务单图号: {summary.total_queries}\n"
                f"结果行数: {summary.total_results}\n"
                f"未命中: {summary.not_found_count}\n"
                f"结果文件: {summary.output_path}"
            ),
        )

    def _on_search_failed(self, exc: Exception, detail: str) -> None:
        self.run_button.config(state="normal")
        self.log("执行失败：")
        self.log(detail)
        
        if isinstance(exc, PermissionError):
            file_name = getattr(exc, 'filename', '目标文件')
            msg = f"文件保存失败！\n\n请检查您是否正在 Excel 中打开了：\n{file_name}\n\n如果是，请先关闭该文件然后再重新搜索。"
            self.messagebox.showerror("文件被占用", msg)
        else:
            self.messagebox.showerror("执行失败", str(exc))


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="BOM 搜索器")
    parser.add_argument("--input", help="待查 Excel 路径")
    parser.add_argument("--bom-folder", help="BOM Excel 文件夹")
    parser.add_argument("--output", help="结果 Excel 路径")
    return parser.parse_args()


def configure_stdout_encoding() -> None:
    try:
        if hasattr(sys.stdout, "reconfigure"):
            sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        if hasattr(sys.stderr, "reconfigure"):
            sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass


def import_tk_modules() -> tuple[Any, Any, Any, Any]:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk

    return tk, ttk, filedialog, messagebox


def main() -> None:
    configure_stdout_encoding()
    args = parse_args()
    if args.input and args.bom_folder and args.output:
        summary = run_search(
            input_file=Path(args.input),
            bom_folder=Path(args.bom_folder),
            output_file=Path(args.output),
            log=print,
        )
        print(
            f"完成: 待查图号 {summary.total_queries} 条，"
            f"结果行 {summary.total_results} 条，"
            f"未命中 {summary.not_found_count} 条。"
        )
        return

    try:
        tk, ttk, filedialog, messagebox = import_tk_modules()
    except Exception as exc:
        raise RuntimeError(
            "当前环境不可用图形界面，请安装/启用 tkinter，或改用命令行参数 --input --bom-folder --output 运行。"
        ) from exc

    root = tk.Tk()
    BOMSearcherApp(root, tk, ttk, filedialog, messagebox)
    root.mainloop()


if __name__ == "__main__":
    main()
