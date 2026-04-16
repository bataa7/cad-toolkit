from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from bom_searcher import RowRecord, extract_children, run_search


def create_workbook(path: Path, rows: list[list[object]], sheet_name: str = "Sheet1") -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for row in rows:
        sheet.append(row)
    workbook.save(path)


class BOMSearcherTests(unittest.TestCase):
    def assert_blank_cell(self, value: object) -> None:
        self.assertIn(value, ("", None))

    def row_map(self, header_row: tuple[object, ...], data_row: tuple[object, ...]) -> dict[str, object]:
        return {
            str(header): data_row[index]
            for index, header in enumerate(header_row)
            if header is not None and str(header).strip()
        }

    def test_blank_sequence_match_does_not_consume_following_sibling(self) -> None:
        rows = [
            RowRecord(55, "55", "", "101002007360", "101002007360", "管夹体", "组件", "组件", "1", "11", "260111"),
            RowRecord(56, "", "", "101002007361", "101002007361", "夹板", "T3", "06Cr19Ni10", "1", "30", "11+19"),
            RowRecord(57, "", "", "101002007362", "101002007362", "棒", "18", "06Cr19Ni10", "1", "11", "L=124mm"),
            RowRecord(58, "56", "", "101002008047", "101002008047", "槽体焊接件", "组件", "组件", "1", "11", "251212"),
        ]

        self.assertEqual(extract_children(rows, 1), [])

    def test_extracts_children_and_blank_sequence_rows(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_file = root / "input.xlsx"
            bom_dir = root / "boms"
            bom_dir.mkdir()
            output_file = root / "result.xlsx"

            create_workbook(
                input_file,
                [
                    ["序号", "图号", "名称", "数量"],
                    [8, "101002000250", "支架总成", 18],
                ],
            )

            create_workbook(
                bom_dir / "bom_1.xlsx",
                [
                    ["序号", "物料ID", "图号", "名称", "厚度", "材质", "数量", "总数量", "备注"],
                    [1, "", "101002000250（248010310）", "支架", "组件", "组件", 1, 1, "260118"],
                    ["1.1", "", "101002000251-A", "弯板", "T3", "06Cr19Ni10", 1, 1, ""],
                    ["1.2", "", "101002000251-B", "弯板", "T3", "06Cr19Ni10", 2, 2, ""],
                    ["", "", "", "工艺说明", "", "", "", "", "左右对称"],
                    [2, "", "101002000293", "角板", "T4", "06Cr19Ni10", 1, 2, "260118"],
                ],
                sheet_name="BOM",
            )

            summary = run_search(input_file, bom_dir, output_file)

            self.assertEqual(summary.total_queries, 1)
            self.assertEqual(summary.total_results, 4)
            self.assertEqual(summary.not_found_count, 0)

            result_book = load_workbook(output_file, data_only=True)
            try:
                result_sheet = result_book["搜索结果"]
                rows = list(result_sheet.iter_rows(values_only=True))
                self.assertEqual(len(rows), 5)
                parent_row = self.row_map(rows[0], rows[1])
                child_1_row = self.row_map(rows[0], rows[2])
                child_2_row = self.row_map(rows[0], rows[3])
                note_row = self.row_map(rows[0], rows[4])
                self.assertEqual(parent_row["序号"], 1)
                self.assertEqual(parent_row["图号"], "101002000250（248010310）")
                self.assertEqual(parent_row["总数量"], 1)
                self.assertEqual(child_1_row["序号"], "1.1")
                self.assertEqual(child_1_row["图号"], "101002000251-A")
                self.assertEqual(child_1_row["总数量"], 1)
                self.assertEqual(child_2_row["序号"], "1.2")
                self.assertEqual(child_2_row["图号"], "101002000251-B")
                self.assertEqual(child_2_row["总数量"], 2)
                self.assert_blank_cell(note_row["序号"])
                self.assertEqual(note_row["名称"], "工艺说明")
                self.assert_blank_cell(note_row["总数量"])
                not_found_sheet = result_book["未找到"]
                not_found_rows = list(not_found_sheet.iter_rows(values_only=True))
                self.assertEqual(len(not_found_rows), 1)
                detail_sheet = result_book["搜索明细"]
                detail_rows = list(detail_sheet.iter_rows(values_only=True))
                self.assertEqual(len(detail_rows), 4)
                self.assertEqual(detail_rows[1][21], "已提取子项")
            finally:
                result_book.close()

    def test_dotted_component_stops_at_next_explicit_sequence_block(self) -> None:
        rows = [
            RowRecord(1, "18.2", "", "COMP-182", "COMP-182", "Sub Assembly", "\u7ec4\u4ef6", "\u7ec4\u4ef6", "1", "1", ""),
            RowRecord(2, "18.2.1", "", "PART-A", "PART-A", "Part A", "T2", "SUS304", "1", "1", ""),
            RowRecord(3, "18.2.2", "", "PART-B", "PART-B", "Part B", "T3", "SUS304", "1", "1", ""),
            RowRecord(4, "1.1", "", "OTHER-COMP", "OTHER-COMP", "Other Sub Assembly", "\u7ec4\u4ef6", "\u7ec4\u4ef6", "1", "1", ""),
            RowRecord(5, "1.1.1", "", "PART-C", "PART-C", "Part C", "T4", "Q235", "1", "1", ""),
            RowRecord(6, "19", "", "TOP-19", "TOP-19", "Next Top Assembly", "\u7ec4\u4ef6", "\u7ec4\u4ef6", "1", "1", ""),
        ]

        children = extract_children(rows, 0)

        self.assertEqual([row.sequence for row in children], ["18.2.1", "18.2.2"])
        self.assertEqual([row.drawing for row in children], ["PART-A", "PART-B"])

    def test_groups_non_dotted_rows_until_next_natural_sequence(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_file = root / "input.xlsx"
            bom_dir = root / "boms"
            bom_dir.mkdir()
            output_file = root / "result.xlsx"

            create_workbook(
                input_file,
                [
                    ["序号", "图号", "名称", "数量"],
                    [55, "101002000878", "手柄总成", 9],
                ],
            )

            create_workbook(
                bom_dir / "bom_2.xlsx",
                [
                    ["序号", "图号", "名称", "厚度", "材质", "数量", "总数量", "备注"],
                    [1, "101002000878（248033200-2）", "手柄", "组件", "组件", 1, 1, ""],
                    ["A", "101002009999", "无1.1格式零件", "T2", "Q235", 1, 1, "特殊件"],
                    ["", "", "工艺说明", "", "", "", "", "同组备注"],
                    [2, "101002001053", "罩", "", "", 1, 1, ""],
                ],
                sheet_name="BOM",
            )

            summary = run_search(input_file, bom_dir, output_file)

            self.assertEqual(summary.total_queries, 1)
            self.assertEqual(summary.total_results, 3)
            self.assertEqual(summary.not_found_count, 0)

            result_book = load_workbook(output_file, data_only=True)
            try:
                result_sheet = result_book["搜索结果"]
                rows = list(result_sheet.iter_rows(values_only=True))
                self.assertEqual(len(rows), 4)
                parent_row = self.row_map(rows[0], rows[1])
                child_row = self.row_map(rows[0], rows[2])
                note_row = self.row_map(rows[0], rows[3])
                self.assertEqual(parent_row["序号"], 1)
                self.assertEqual(parent_row["图号"], "101002000878（248033200-2）")
                self.assertEqual(parent_row["总数量"], 1)
                self.assertEqual(child_row["序号"], "A")
                self.assertEqual(child_row["图号"], "101002009999")
                self.assertEqual(child_row["总数量"], 1)
                self.assert_blank_cell(note_row["序号"])
                self.assertEqual(note_row["名称"], "工艺说明")
                self.assert_blank_cell(note_row["总数量"])
                self.assertEqual(note_row["备注"], "同组备注")
            finally:
                result_book.close()

    def test_includes_parent_when_match_has_no_children(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_file = root / "input.xlsx"
            bom_dir = root / "boms"
            bom_dir.mkdir()
            output_file = root / "result.xlsx"

            create_workbook(
                input_file,
                [
                    ["序号", "图号", "名称", "数量"],
                    [123, "101002000670", "方钢", 36],
                ],
            )

            create_workbook(
                bom_dir / "bom_3.xlsx",
                [
                    ["序号", "图号", "名称", "厚度", "材质", "数量", "总数量", "备注"],
                    [99, "101002000670（248031200-2）", "方钢", "", "", 1, 36, ""],
                    [100, "101002000671", "板", "", "", 1, 44, ""],
                ],
                sheet_name="BOM",
            )

            summary = run_search(input_file, bom_dir, output_file)

            self.assertEqual(summary.total_queries, 1)
            self.assertEqual(summary.total_results, 1)
            self.assertEqual(summary.not_found_count, 0)

            result_book = load_workbook(output_file, data_only=True)
            try:
                result_sheet = result_book["搜索结果"]
                rows = list(result_sheet.iter_rows(values_only=True))
                self.assertEqual(len(rows), 2)
                parent_row = self.row_map(rows[0], rows[1])
                self.assertEqual(parent_row["序号"], 99)
                self.assertEqual(parent_row["图号"], "101002000670（248031200-2）")
                self.assertEqual(parent_row["总数量"], 36)

                detail_sheet = result_book["搜索明细"]
                detail_rows = list(detail_sheet.iter_rows(values_only=True))
                self.assertEqual(len(detail_rows), 2)
                self.assertEqual(detail_rows[1][10], "101002000670(248031200-2)")
                self.assertEqual(detail_rows[1][21], "仅父项，无子项")

                not_found_sheet = result_book["未找到"]
                not_found_rows = list(not_found_sheet.iter_rows(values_only=True))
                self.assertEqual(len(not_found_rows), 1)
            finally:
                result_book.close()

    def test_prefers_explicit_sequence_match_over_blank_sequence_usage_row(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_file = root / "input.xlsx"
            bom_dir = root / "boms"
            bom_dir.mkdir()
            output_file = root / "result.xlsx"

            create_workbook(
                input_file,
                [
                    ["序号", "图号", "名称", "数量"],
                    [157, "101002007361", "夹板", 12],
                ],
            )

            create_workbook(
                bom_dir / "bom_explicit.xlsx",
                [
                    ["序号", "图号", "名称", "厚度", "材质", "数量", "总数量", "备注"],
                    [55, "101002007360", "管夹体", "组件", "组件", 1, 11, "260111"],
                    ["", "101002007361", "夹板", "T3", "06Cr19Ni10", 1, 30, "11+19"],
                    ["", "101002007362", "棒", 18, "06Cr19Ni10", 1, 11, "L=124mm"],
                    [56, "101002007361", "夹板", "T3", "06Cr19Ni10", 1, "", "260111"],
                ],
                sheet_name="BOM",
            )

            summary = run_search(input_file, bom_dir, output_file)

            self.assertEqual(summary.total_queries, 1)
            self.assertEqual(summary.total_results, 1)
            self.assertEqual(summary.not_found_count, 0)

            result_book = load_workbook(output_file, data_only=True)
            try:
                result_sheet = result_book["搜索结果"]
                rows = list(result_sheet.iter_rows(values_only=True))
                self.assertEqual(len(rows), 2)
                parent_row = self.row_map(rows[0], rows[1])
                self.assertEqual(parent_row["序号"], 56)
                self.assertEqual(parent_row["图号"], "101002007361")
                self.assertEqual(parent_row["名称"], "夹板")
                self.assert_blank_cell(parent_row["总数量"])
                self.assertEqual(parent_row["备注"], "260111")
            finally:
                result_book.close()

    def test_prefers_natural_sequence_match_over_dotted_sequence_match(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_file = root / "input.xlsx"
            bom_dir = root / "boms"
            bom_dir.mkdir()
            output_file = root / "result.xlsx"

            create_workbook(
                input_file,
                [
                    ["序号", "图号", "名称", "数量"],
                    [133, "101002004702", "夹板", 252],
                ],
            )

            create_workbook(
                bom_dir / "bom_natural_and_dotted.xlsx",
                [
                    ["序号", "图号", "名称", "厚度", "材质", "数量", "总数量", "备注"],
                    [35, "101002004702", "夹板", "T3", "06Cr19Ni10", 1, 242, "251212 121+121"],
                    [44, "101002006634", "管夹体", "组件", "组件", 1, 121, "251212"],
                    ["44.1", "101002004702", "夹板", "T3", "06Cr19Ni10", 1, "", ""],
                    ["44.2", "101002006633", "棒", "/", "06Cr19Ni10", 1, 121, "φ18圆钢 L=120mm外协"],
                ],
                sheet_name="BOM",
            )

            summary = run_search(input_file, bom_dir, output_file)

            self.assertEqual(summary.total_queries, 1)
            self.assertEqual(summary.total_results, 1)
            self.assertEqual(summary.not_found_count, 0)

            result_book = load_workbook(output_file, data_only=True)
            try:
                result_sheet = result_book["搜索结果"]
                rows = list(result_sheet.iter_rows(values_only=True))
                self.assertEqual(len(rows), 2)
                parent_row = self.row_map(rows[0], rows[1])
                self.assertEqual(parent_row["序号"], 35)
                self.assertEqual(parent_row["图号"], "101002004702")
                self.assertEqual(parent_row["名称"], "夹板")
                self.assertEqual(parent_row["总数量"], 242)
                self.assertEqual(parent_row["备注"], "251212 121+121")
            finally:
                result_book.close()

    def test_search_result_preserves_original_bom_extra_columns(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_file = root / "input.xlsx"
            bom_dir = root / "boms"
            bom_dir.mkdir()
            output_file = root / "result.xlsx"

            create_workbook(
                input_file,
                [
                    ["序号", "图号", "名称", "数量"],
                    [157, "101002007361", "夹板", 12],
                ],
            )

            create_workbook(
                bom_dir / "bom_raw_columns.xlsx",
                [
                    ["序号", "物料ID", "图号", "名称", "厚度", "材质", "数量", "总数量", "备注", "扩展1", "扩展2"],
                    [56, "", "101002007361", "夹板", "T3", "06Cr19Ni10", 1, "", "260111", "J值", 123.45],
                ],
                sheet_name="BOM",
            )

            run_search(input_file, bom_dir, output_file)

            result_book = load_workbook(output_file, data_only=True)
            try:
                result_sheet = result_book["搜索结果"]
                rows = list(result_sheet.iter_rows(values_only=True))
                self.assertEqual(rows[0][9], "扩展1")
                self.assertEqual(rows[0][10], "扩展2")
                self.assertEqual(rows[1][0], 56)
                self.assertEqual(rows[1][8], "260111")
                self.assertEqual(rows[1][9], "J值")
                self.assertEqual(rows[1][10], 123.45)
            finally:
                result_book.close()

    def test_search_result_aligns_rows_when_material_id_column_is_missing_in_some_sources(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_file = root / "input.xlsx"
            bom_dir = root / "boms"
            bom_dir.mkdir()
            output_file = root / "result.xlsx"

            create_workbook(
                input_file,
                [
                    ["序号", "图号", "名称", "数量"],
                    [1, "PART-A", "零件A", 2],
                    [2, "PART-B", "零件B", 3],
                ],
            )

            create_workbook(
                bom_dir / "with_material_id.xlsx",
                [
                    ["序号", "物料ID", "图号", "名称", "厚度", "材质", "数量", "总数量", "备注"],
                    [11, "MAT-A", "PART-A", "零件A", "T2", "SUS304", 1, 2, "A备注"],
                ],
                sheet_name="BOM",
            )

            create_workbook(
                bom_dir / "without_material_id.xlsx",
                [
                    ["序号", "图号", "名称", "厚度", "材质", "数量", "总数量", "备注"],
                    [22, "PART-B", "零件B", "T5", "Q235", 1, 3, "B备注"],
                ],
                sheet_name="BOM",
            )

            run_search(input_file, bom_dir, output_file)

            result_book = load_workbook(output_file, data_only=True)
            try:
                result_sheet = result_book["搜索结果"]
                rows = list(result_sheet.iter_rows(values_only=True))
                self.assertEqual(rows[0], ("序号", "物料ID", "图号", "名称", "厚度", "材质", "数量", "总数量", "备注"))

                first_row = self.row_map(rows[0], rows[1])
                second_row = self.row_map(rows[0], rows[2])

                self.assertEqual(first_row["物料ID"], "MAT-A")
                self.assertEqual(first_row["图号"], "PART-A")
                self.assertEqual(first_row["厚度"], "T2")
                self.assertEqual(first_row["备注"], "A备注")

                self.assert_blank_cell(second_row["物料ID"])
                self.assertEqual(second_row["图号"], "PART-B")
                self.assertEqual(second_row["名称"], "零件B")
                self.assertEqual(second_row["厚度"], "T5")
                self.assertEqual(second_row["材质"], "Q235")
                self.assertEqual(second_row["备注"], "B备注")
            finally:
                result_book.close()

    def test_search_result_maps_spec_and_total_alias_headers(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_file = root / "input.xlsx"
            bom_dir = root / "boms"
            bom_dir.mkdir()
            output_file = root / "result.xlsx"

            create_workbook(
                input_file,
                [
                    ["序号", "图号", "名称", "数量"],
                    [1, "PART-C", "零件C", 5],
                ],
            )

            create_workbook(
                bom_dir / "alias_headers.xlsx",
                [
                    ["序号", "图号", "名称", "规格", "材质", "数量", "总量", "备注"],
                    [31, "PART-C", "零件C", "T8", "Q235", 1, 9, "别名表头"],
                ],
                sheet_name="BOM",
            )

            run_search(input_file, bom_dir, output_file)

            result_book = load_workbook(output_file, data_only=True)
            try:
                result_sheet = result_book["搜索结果"]
                rows = list(result_sheet.iter_rows(values_only=True))
                row = self.row_map(rows[0], rows[1])
                self.assertEqual(row["图号"], "PART-C")
                self.assertEqual(row["厚度"], "T8")
                self.assertEqual(row["总数量"], 9)
                self.assertEqual(row["备注"], "别名表头")
            finally:
                result_book.close()

    def test_search_result_ignores_sparse_far_right_tail_cells(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_file = root / "input.xlsx"
            bom_dir = root / "boms"
            bom_dir.mkdir()
            output_file = root / "result.xlsx"

            create_workbook(
                input_file,
                [
                    ["序号", "图号", "名称", "数量"],
                    [98, "164090500", "支座", 2],
                ],
            )

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "生产BOM"
            sheet.append(["序号", "物料ID", "图号", "名称", "厚度", "材质", "数量", "总数量", "备注"])
            sheet.append([98, "", "164090500", "支座", "组件", "组件", 1, 2, "250505"])
            sheet.cell(row=2, column=16384, value=164341106)
            workbook.save(bom_dir / "sparse_tail.xlsx")
            workbook.close()

            summary = run_search(input_file, bom_dir, output_file)

            self.assertEqual(summary.total_queries, 1)
            self.assertEqual(summary.total_results, 1)
            self.assertEqual(summary.not_found_count, 0)

            result_book = load_workbook(output_file, data_only=True)
            try:
                result_sheet = result_book["搜索结果"]
                rows = list(result_sheet.iter_rows(values_only=True))
                self.assertLessEqual(result_sheet.max_column, 9)
                self.assertEqual(rows[1][0], 98)
                self.assertEqual(rows[1][2], "164090500")
            finally:
                result_book.close()

    def test_prefers_latest_source_when_same_drawing_hits_multiple_boms(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_file = root / "input.xlsx"
            bom_dir = root / "boms"
            bom_dir.mkdir()
            output_file = root / "result.xlsx"

            create_workbook(
                input_file,
                [
                    ["序号", "图号", "名称", "数量"],
                    [6, "101002007076", "护板", 6],
                    [233, "101002007076", "护板", 3],
                ],
            )

            create_workbook(
                bom_dir / "250907_old.xlsx",
                [
                    ["序号", "图号", "名称", "厚度", "材质", "数量", "总数量", "备注"],
                    [10, "101002007076", "护板", "T2.5", "06Cr19Ni10", 1, 2, "250907"],
                ],
                sheet_name="BOM",
            )

            create_workbook(
                bom_dir / "260118_new.xlsx",
                [
                    ["序号", "图号", "名称", "厚度", "材质", "数量", "总数量", "备注"],
                    [20, "101002007076", "护板", "T2.5", "06Cr19Ni10", 1, 11, "260118"],
                ],
                sheet_name="BOM",
            )

            summary = run_search(input_file, bom_dir, output_file)

            self.assertEqual(summary.total_queries, 2)
            self.assertEqual(summary.total_results, 2)
            self.assertEqual(summary.not_found_count, 0)

            result_book = load_workbook(output_file, data_only=True)
            try:
                result_sheet = result_book["搜索结果"]
                rows = list(result_sheet.iter_rows(values_only=True))
                self.assertEqual(len(rows), 3)
                first_row = self.row_map(rows[0], rows[1])
                second_row = self.row_map(rows[0], rows[2])
                self.assertEqual(first_row["序号"], 20)
                self.assertEqual(first_row["备注"], "260118")
                self.assertEqual(second_row["序号"], 20)
                self.assertEqual(second_row["备注"], "260118")

                detail_sheet = result_book["搜索明细"]
                detail_rows = list(detail_sheet.iter_rows(values_only=True))
                self.assertEqual(len(detail_rows), 3)
                self.assertEqual(detail_rows[1][6], "260118_new.xlsx")
                self.assertEqual(detail_rows[2][6], "260118_new.xlsx")
            finally:
                result_book.close()

    def test_ignores_diff_workbooks_and_report_sheets_during_search(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_file = root / "input.xlsx"
            bom_dir = root / "boms"
            bom_dir.mkdir()
            output_file = root / "result.xlsx"

            create_workbook(
                input_file,
                [
                    ["序号", "图号", "名称", "数量"],
                    [18, "100-A", "铰链", 24],
                ],
            )

            create_workbook(
                bom_dir / "260313BOM.xlsx",
                [
                    ["序号", "图号", "名称", "厚度", "材质", "数量", "总数量", "备注"],
                    [18, "100-A", "铰链", "组件", "组件", 1, 24, "250515"],
                    ["18.1", "PART-001", "管", "/", "06Cr19Ni10", 1, "", ""],
                    ["18.2", "PART-002", "板", "T8", "06Cr19Ni10", 1, "", ""],
                    [19, "NEXT-001", "把手", "/", "06Cr19Ni10", 1, 8, ""],
                ],
                sheet_name="BOM",
            )

            create_workbook(
                bom_dir / "260313BOM_vs_search_diff.xlsx",
                [
                    ["序号", "图号", "名称", "厚度", "材质", "数量", "总数量", "备注"],
                    [18, "100-A", "铰链", "组件", "组件", 1, 24, "250515"],
                    ["18.1", "PART-001", "管", "/", "06Cr19Ni10", 1, "", ""],
                    ["18.2", "PART-002", "板", "T8", "06Cr19Ni10", 1, "", ""],
                    ["80.2.5", "PART-EXTRA-1", "变形管", "组件", "组件", 1, "", ""],
                    ["80.2.5.1", "PART-EXTRA-2", "弯板", "T2", "06Cr19Ni10", 2, "", ""],
                    ["80.2.5.2", "PART-EXTRA-3", "法兰", "T10", "06Cr19Ni10", 1, "", ""],
                ],
                sheet_name="OnlyManualSeq",
            )

            summary = run_search(input_file, bom_dir, output_file)

            self.assertEqual(summary.total_queries, 1)
            self.assertEqual(summary.total_results, 3)
            self.assertEqual(summary.not_found_count, 0)

            result_book = load_workbook(output_file, data_only=True)
            try:
                result_sheet = result_book["搜索结果"]
                rows = list(result_sheet.iter_rows(values_only=True))
                self.assertEqual(len(rows), 4)
                parent_row = self.row_map(rows[0], rows[1])
                child_1_row = self.row_map(rows[0], rows[2])
                child_2_row = self.row_map(rows[0], rows[3])
                self.assertEqual(parent_row["序号"], 18)
                self.assertEqual(parent_row["图号"], "100-A")
                self.assertEqual(child_1_row["序号"], "18.1")
                self.assertEqual(child_1_row["图号"], "PART-001")
                self.assertEqual(child_2_row["序号"], "18.2")
                self.assertEqual(child_2_row["图号"], "PART-002")

                detail_sheet = result_book["搜索明细"]
                detail_rows = list(detail_sheet.iter_rows(values_only=True))
                self.assertEqual(len(detail_rows), 3)
                self.assertEqual(detail_rows[1][6], "260313BOM.xlsx")
                self.assertEqual(detail_rows[2][6], "260313BOM.xlsx")
            finally:
                result_book.close()

    def test_prefers_latest_file_version_over_older_row_remark(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_file = root / "input.xlsx"
            bom_dir = root / "boms"
            bom_dir.mkdir()
            output_file = root / "result.xlsx"

            create_workbook(
                input_file,
                [
                    ["\u5e8f\u53f7", "\u56fe\u53f7", "\u540d\u79f0", "\u6570\u91cf"],
                    [7, "PART-001", "\u677f", 1],
                ],
            )

            create_workbook(
                bom_dir / "250515_old.xlsx",
                [
                    ["\u5e8f\u53f7", "\u56fe\u53f7", "\u540d\u79f0", "\u539a\u5ea6", "\u6750\u8d28", "\u6570\u91cf", "\u603b\u6570\u91cf", "\u5907\u6ce8"],
                    [10, "PART-001", "\u677f", "T5", "022Cr17Ni12Mo2", 1, 2, "250221"],
                ],
                sheet_name="BOM",
            )

            create_workbook(
                bom_dir / "251110_new.xlsx",
                [
                    ["\u5e8f\u53f7", "\u56fe\u53f7", "\u540d\u79f0", "\u539a\u5ea6", "\u6750\u8d28", "\u6570\u91cf", "\u603b\u6570\u91cf", "\u5907\u6ce8"],
                    [20, "PART-001", "\u677f", "T5", "06Cr19Ni10", 1, 2, ""],
                ],
                sheet_name="BOM",
            )

            run_search(input_file, bom_dir, output_file)

            result_book = load_workbook(output_file, data_only=True)
            try:
                result_sheet = result_book["搜索结果"]
                rows = list(result_sheet.iter_rows(values_only=True))
                parent_row = self.row_map(rows[0], rows[1])
                self.assertEqual(parent_row["序号"], 20)
                self.assertEqual(parent_row["材质"], "06Cr19Ni10")

                detail_sheet = result_book["搜索明细"]
                detail_rows = list(detail_sheet.iter_rows(values_only=True))
                self.assertEqual(detail_rows[1][6], "251110_new.xlsx")
            finally:
                result_book.close()

    def test_prefers_closer_total_quantity_before_row_remark_when_file_versions_tie(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_file = root / "input.xlsx"
            bom_dir = root / "boms"
            bom_dir.mkdir()
            output_file = root / "result.xlsx"

            create_workbook(
                input_file,
                [
                    ["\u5e8f\u53f7", "\u56fe\u53f7", "\u540d\u79f0", "\u6570\u91cf"],
                    [19, "PART-002", "\u628a\u624b", 6],
                ],
            )

            create_workbook(
                bom_dir / "251110_old_remark.xlsx",
                [
                    ["\u5e8f\u53f7", "\u56fe\u53f7", "\u540d\u79f0", "\u539a\u5ea6", "\u6750\u8d28", "\u6570\u91cf", "\u603b\u6570\u91cf", "\u5907\u6ce8"],
                    [30, "PART-002", "\u628a\u624b", "/", "304", 2, 4, "260201"],
                ],
                sheet_name="BOM",
            )

            create_workbook(
                bom_dir / "251110_quantity_match.xlsx",
                [
                    ["\u5e8f\u53f7", "\u56fe\u53f7", "\u540d\u79f0", "\u539a\u5ea6", "\u6750\u8d28", "\u6570\u91cf", "\u603b\u6570\u91cf", "\u5907\u6ce8"],
                    [40, "PART-002", "\u628a\u624b", "/", "304", 1, 6, ""],
                ],
                sheet_name="BOM",
            )

            run_search(input_file, bom_dir, output_file)

            result_book = load_workbook(output_file, data_only=True)
            try:
                result_sheet = result_book["搜索结果"]
                rows = list(result_sheet.iter_rows(values_only=True))
                parent_row = self.row_map(rows[0], rows[1])
                self.assertEqual(parent_row["序号"], 40)
                self.assertEqual(parent_row["数量"], 1)
                self.assertEqual(parent_row["总数量"], 6)

                detail_sheet = result_book["搜索明细"]
                detail_rows = list(detail_sheet.iter_rows(values_only=True))
                self.assertEqual(detail_rows[1][6], "251110_quantity_match.xlsx")
            finally:
                result_book.close()

    def test_parent_quantity_is_normalized_to_one_in_display_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_file = root / "input.xlsx"
            bom_dir = root / "boms"
            bom_dir.mkdir()
            output_file = root / "result.xlsx"

            create_workbook(
                input_file,
                [
                    ["\u5e8f\u53f7", "\u56fe\u53f7", "\u540d\u79f0", "\u6570\u91cf"],
                    [19, "PART-003", "\u628a\u624b", 8],
                ],
            )

            create_workbook(
                bom_dir / "part.xlsx",
                [
                    ["\u5e8f\u53f7", "\u56fe\u53f7", "\u540d\u79f0", "\u539a\u5ea6", "\u6750\u8d28", "\u6570\u91cf", "\u603b\u6570\u91cf", "\u5907\u6ce8"],
                    [30, "PART-003", "\u628a\u624b", "/", "304", 2, 4, ""],
                ],
                sheet_name="BOM",
            )

            run_search(input_file, bom_dir, output_file)

            result_book = load_workbook(output_file, data_only=True)
            try:
                aggregate_sheet = result_book["汇总BOM"]
                rows = list(aggregate_sheet.iter_rows(values_only=True))
                self.assertEqual(rows[1][0], "19")
                self.assertEqual(rows[1][6], "1")
                self.assertEqual(rows[1][7], "8")
            finally:
                result_book.close()

    def test_reports_not_found_when_no_match_exists(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_file = root / "input.xlsx"
            bom_dir = root / "boms"
            bom_dir.mkdir()
            output_file = root / "result.xlsx"

            create_workbook(
                input_file,
                [
                    ["序号", "图号", "名称", "数量"],
                    [1, "999999999999", "不存在的总成", 1],
                ],
            )

            create_workbook(
                bom_dir / "bom_1.xlsx",
                [
                    ["序号", "图号", "名称", "数量"],
                    [1, "101002000250", "支架", 1],
                    ["1.1", "101002000251", "弯板", 1],
                ],
            )

            summary = run_search(input_file, bom_dir, output_file)

            self.assertEqual(summary.total_results, 0)
            self.assertEqual(summary.not_found_count, 1)

            result_book = load_workbook(output_file, data_only=True)
            try:
                not_found_sheet = result_book["未找到"]
                rows = list(not_found_sheet.iter_rows(values_only=True))
                self.assertEqual(rows[1][3], "999999999999")
                self.assertIn("未找到匹配图号", rows[1][6])
            finally:
                result_book.close()

    def test_summary_sheet_uses_display_bom_content_and_a4_landscape(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_file = root / "input.xlsx"
            bom_dir = root / "boms"
            bom_dir.mkdir()
            output_file = root / "result.xlsx"

            create_workbook(
                input_file,
                [
                    ["序号", "图号", "名称", "数量"],
                    [10, "100-A", "总成A", 2],
                    [20, "100-B", "总成B", 3],
                ],
            )

            create_workbook(
                bom_dir / "bom_a.xlsx",
                [
                    ["序号", "物料ID", "图号", "名称", "厚度", "材质", "数量", "总数量", "备注"],
                    [1, "PARENT-A", "100-A", "总成A", "组件", "组件", 1, 1, ""],
                    ["1.1", "CH-001", "PART-X", "子件X", "T2", "SUS304", 2, 2, "共用"],
                    ["1.2", "CH-002", "PART-Y", "子件Y", "T1", "Q235", 1, 1, ""],
                ],
                sheet_name="BOM",
            )

            create_workbook(
                bom_dir / "bom_b.xlsx",
                [
                    ["序号", "物料ID", "图号", "名称", "厚度", "材质", "数量", "总数量", "备注"],
                    [1, "PARENT-B", "100-B", "总成B", "组件", "组件", 1, 1, ""],
                    ["1.1", "CH-001", "PART-X", "子件X", "T2", "SUS304", 2, 2, "共用"],
                    ["1.2", "CH-003", "PART-Z", "子件Z", "T3", "AL", 4, 4, ""],
                ],
                sheet_name="BOM",
            )

            summary = run_search(input_file, bom_dir, output_file)

            self.assertEqual(summary.total_queries, 2)
            self.assertEqual(summary.not_found_count, 0)

            result_book = load_workbook(output_file, data_only=True)
            try:
                aggregate_sheet = result_book["汇总BOM"]
                rows = list(aggregate_sheet.iter_rows(values_only=True))
                self.assertEqual(rows[0], ("序号", "物料ID", "图号", "名称", "厚度", "材质", "数量", "总数量", "备注"))
                self.assertEqual(len(rows), 7)
                parent_a_row = self.row_map(rows[0], rows[1])
                part_x_row = self.row_map(rows[0], rows[2])
                part_y_row = self.row_map(rows[0], rows[3])
                parent_b_row = self.row_map(rows[0], rows[4])
                second_part_x_row = self.row_map(rows[0], rows[5])
                part_z_row = self.row_map(rows[0], rows[6])
                self.assertEqual(parent_a_row["序号"], "10")
                self.assertEqual(parent_a_row["物料ID"], "PARENT-A")
                self.assertEqual(parent_a_row["图号"], "100-A")
                self.assertEqual(parent_a_row["数量"], "1")
                self.assertEqual(parent_a_row["总数量"], "2")
                self.assertEqual(part_x_row["序号"], "10.1")
                self.assertEqual(part_x_row["物料ID"], "CH-001")
                self.assertEqual(part_x_row["图号"], "PART-X")
                self.assertEqual(part_x_row["厚度"], "T2")
                self.assertEqual(part_x_row["材质"], "SUS304")
                self.assertEqual(part_x_row["数量"], "2")
                self.assert_blank_cell(part_x_row["总数量"])
                self.assertEqual(part_x_row["备注"], "共用")
                self.assertEqual(part_y_row["序号"], "10.2")
                self.assertEqual(part_y_row["物料ID"], "CH-002")
                self.assertEqual(part_y_row["图号"], "PART-Y")
                self.assertEqual(part_y_row["数量"], "1")
                self.assert_blank_cell(part_y_row["总数量"])
                self.assertEqual(parent_b_row["序号"], "20")
                self.assertEqual(parent_b_row["物料ID"], "PARENT-B")
                self.assertEqual(parent_b_row["图号"], "100-B")
                self.assertEqual(parent_b_row["数量"], "1")
                self.assertEqual(parent_b_row["总数量"], "3")
                self.assertEqual(second_part_x_row["序号"], "20.1")
                self.assertEqual(second_part_x_row["物料ID"], "CH-001")
                self.assertEqual(second_part_x_row["图号"], "PART-X")
                self.assertEqual(second_part_x_row["数量"], "2")
                self.assert_blank_cell(second_part_x_row["总数量"])
                self.assertEqual(part_z_row["序号"], "20.2")
                self.assertEqual(part_z_row["物料ID"], "CH-003")
                self.assertEqual(part_z_row["图号"], "PART-Z")
                self.assertEqual(part_z_row["数量"], "4")
                self.assert_blank_cell(part_z_row["总数量"])
                self.assertEqual(aggregate_sheet.page_setup.orientation, "landscape")
                self.assertEqual(str(aggregate_sheet.page_setup.paperSize), aggregate_sheet.PAPERSIZE_A4)
            finally:
                result_book.close()

    def test_summary_sheet_follows_task_order_and_keeps_duplicate_queries(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_file = root / "input.xlsx"
            bom_dir = root / "boms"
            bom_dir.mkdir()
            output_file = root / "result.xlsx"

            create_workbook(
                input_file,
                [
                    ["序号", "图号", "名称", "数量"],
                    [10, "100-A", "总成A", 2],
                    [11, "100-A", "总成A", 7],
                ],
            )

            create_workbook(
                bom_dir / "bom_a.xlsx",
                [
                    ["序号", "物料ID", "图号", "名称", "厚度", "材质", "数量", "总数量", "备注"],
                    [1, "PARENT-A", "100-A", "总成A", "组件", "组件", 1, 1, ""],
                    ["1.1", "CH-001", "PART-X", "子件X", "T2", "SUS304", 2, 2, "共用"],
                ],
                sheet_name="BOM",
            )

            run_search(input_file, bom_dir, output_file)

            result_book = load_workbook(output_file, data_only=True)
            try:
                aggregate_sheet = result_book["汇总BOM"]
                rows = list(aggregate_sheet.iter_rows(values_only=True))
                self.assertEqual(len(rows), 5)

                first_parent_row = self.row_map(rows[0], rows[1])
                first_child_row = self.row_map(rows[0], rows[2])
                second_parent_row = self.row_map(rows[0], rows[3])
                second_child_row = self.row_map(rows[0], rows[4])

                self.assertEqual(first_parent_row["序号"], "10")
                self.assertEqual(first_parent_row["图号"], "100-A")
                self.assertEqual(first_parent_row["总数量"], "2")
                self.assertEqual(first_child_row["序号"], "10.1")
                self.assertEqual(first_child_row["图号"], "PART-X")
                self.assert_blank_cell(first_child_row["总数量"])

                self.assertEqual(second_parent_row["序号"], "11")
                self.assertEqual(second_parent_row["图号"], "100-A")
                self.assertEqual(second_parent_row["总数量"], "7")
                self.assertEqual(second_child_row["序号"], "11.1")
                self.assertEqual(second_child_row["图号"], "PART-X")
                self.assert_blank_cell(second_child_row["总数量"])
            finally:
                result_book.close()

    def test_summary_sheet_keeps_component_and_subset_children(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_file = root / "input.xlsx"
            bom_dir = root / "boms"
            bom_dir.mkdir()
            output_file = root / "result.xlsx"

            create_workbook(
                input_file,
                [
                    ["序号", "图号", "名称", "数量"],
                    [10, "100-A", "总成A", 2],
                ],
            )

            create_workbook(
                bom_dir / "bom_nested.xlsx",
                [
                    ["序号", "物料ID", "图号", "名称", "厚度", "材质", "数量", "总数量", "备注"],
                    [1, "PARENT-A", "100-A", "总成A", "组件", "组件", 1, 1, ""],
                    ["1.1", "SUB-001", "SUBSET-A", "子集A", "组件", "组件", 1, 1, ""],
                    ["1.1.1", "LEAF-001", "PART-X", "子件X", "T2", "SUS304", 2, 2, ""],
                    ["1.1.2", "LEAF-002", "PART-Y", "子件Y", "T1", "Q235", 3, 3, ""],
                    ["1.2", "LEAF-003", "PART-Z", "子件Z", "T5", "AL", 1, 1, ""],
                ],
                sheet_name="BOM",
            )

            run_search(input_file, bom_dir, output_file)

            result_book = load_workbook(output_file, data_only=True)
            try:
                aggregate_sheet = result_book["汇总BOM"]
                rows = list(aggregate_sheet.iter_rows(values_only=True))
                self.assertEqual(len(rows), 6)

                parent_row = self.row_map(rows[0], rows[1])
                subset_row = self.row_map(rows[0], rows[2])
                first_row = self.row_map(rows[0], rows[3])
                second_row = self.row_map(rows[0], rows[4])
                third_row = self.row_map(rows[0], rows[5])

                self.assertEqual(parent_row["序号"], "10")
                self.assertEqual(parent_row["图号"], "100-A")
                self.assertEqual(parent_row["总数量"], "2")

                self.assertEqual(subset_row["序号"], "10.1")
                self.assertEqual(subset_row["图号"], "SUBSET-A")
                self.assert_blank_cell(subset_row["总数量"])

                self.assertEqual(first_row["序号"], "10.1.1")
                self.assertEqual(first_row["图号"], "PART-X")
                self.assert_blank_cell(first_row["总数量"])

                self.assertEqual(second_row["序号"], "10.1.2")
                self.assertEqual(second_row["图号"], "PART-Y")
                self.assert_blank_cell(second_row["总数量"])

                self.assertEqual(third_row["序号"], "10.2")
                self.assertEqual(third_row["图号"], "PART-Z")
                self.assert_blank_cell(third_row["总数量"])

                summary_drawings = [row[2] for row in rows[1:]]
                self.assertIn("SUBSET-A", summary_drawings)
            finally:
                result_book.close()

    def test_summary_sheet_skips_subset_parent_fallback(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_file = root / "input.xlsx"
            bom_dir = root / "boms"
            bom_dir.mkdir()
            output_file = root / "result.xlsx"

            create_workbook(
                input_file,
                [
                    ["序号", "图号", "名称", "数量"],
                    ["80.2.5", "SUBSET-A", "子集A", 2],
                ],
            )

            create_workbook(
                bom_dir / "subset_only.xlsx",
                [
                    ["序号", "物料ID", "图号", "名称", "厚度", "材质", "数量", "总数量", "备注"],
                    ["80.2.5", "SUB-001", "SUBSET-A", "子集A", "组件", "组件", 1, 1, ""],
                ],
                sheet_name="BOM",
            )

            run_search(input_file, bom_dir, output_file)

            result_book = load_workbook(output_file, data_only=True)
            try:
                aggregate_sheet = result_book["汇总BOM"]
                rows = list(aggregate_sheet.iter_rows(values_only=True))
                self.assertEqual(len(rows), 2)
                subset_row = self.row_map(rows[0], rows[1])
                self.assertEqual(subset_row["序号"], "80.2.5")
                self.assertEqual(subset_row["图号"], "SUBSET-A")
                self.assertEqual(subset_row["总数量"], "2")
            finally:
                result_book.close()

    def test_summary_sheet_falls_back_to_display_parent_when_no_children(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_file = root / "input.xlsx"
            bom_dir = root / "boms"
            bom_dir.mkdir()
            output_file = root / "result.xlsx"

            create_workbook(
                input_file,
                [
                    ["序号", "图号", "名称", "数量"],
                    [1, "SOLO-100", "单独总成", 5],
                ],
            )

            create_workbook(
                bom_dir / "solo.xlsx",
                [
                    ["序号", "物料ID", "图号", "名称", "厚度", "材质", "数量", "总数量", "备注"],
                    [1, "SOLO-ID", "SOLO-100", "单独总成", "", "", 1, 1, "仅父项"],
                ],
                sheet_name="BOM",
            )

            run_search(input_file, bom_dir, output_file)

            result_book = load_workbook(output_file, data_only=True)
            try:
                aggregate_sheet = result_book["汇总BOM"]
                rows = list(aggregate_sheet.iter_rows(values_only=True))
                self.assertEqual(len(rows), 2)
                solo_row = self.row_map(rows[0], rows[1])
                self.assertEqual(solo_row["序号"], "1")
                self.assertEqual(solo_row["物料ID"], "SOLO-ID")
                self.assertEqual(solo_row["图号"], "SOLO-100")
                self.assertEqual(solo_row["名称"], "单独总成")
                self.assertEqual(solo_row["数量"], "1")
                self.assertEqual(solo_row["总数量"], "5")
                self.assertEqual(solo_row["备注"], "仅父项")
            finally:
                result_book.close()

    def test_summary_sheet_keeps_blank_thickness_as_is(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_file = root / "input.xlsx"
            bom_dir = root / "boms"
            bom_dir.mkdir()
            output_file = root / "result.xlsx"

            create_workbook(
                input_file,
                [
                    ["序号", "图号", "名称", "数量"],
                    [10, "100-A", "总成A", 2],
                    [20, "100-B", "总成B", 3],
                ],
            )

            create_workbook(
                bom_dir / "bom_a.xlsx",
                [
                    ["序号", "物料ID", "图号", "名称", "厚度", "材质", "数量", "总数量", "备注"],
                    [1, "PARENT-A", "100-A", "总成A", "组件", "组件", 1, 1, ""],
                    ["1.1", "CH-001", "PART-X", "子件X", "", "SUS304", 2, 2, ""],
                ],
                sheet_name="BOM",
            )

            create_workbook(
                bom_dir / "bom_b.xlsx",
                [
                    ["序号", "物料ID", "图号", "名称", "厚度", "材质", "数量", "总数量", "备注"],
                    [1, "PARENT-B", "100-B", "总成B", "组件", "组件", 1, 1, ""],
                    ["1.1", "CH-001", "PART-X", "子件X", "T2", "SUS304", 2, 2, ""],
                ],
                sheet_name="BOM",
            )

            run_search(input_file, bom_dir, output_file)

            result_book = load_workbook(output_file, data_only=True)
            try:
                aggregate_sheet = result_book["汇总BOM"]
                rows = list(aggregate_sheet.iter_rows(values_only=True))
                self.assertEqual(len(rows), 5)
                first_parent_row = self.row_map(rows[0], rows[1])
                first_row = self.row_map(rows[0], rows[2])
                second_parent_row = self.row_map(rows[0], rows[3])
                second_row = self.row_map(rows[0], rows[4])

                self.assertEqual(first_parent_row["图号"], "100-A")
                self.assertEqual(first_parent_row["总数量"], "2")

                self.assertEqual(first_row["物料ID"], "CH-001")
                self.assertEqual(first_row["图号"], "PART-X")
                self.assert_blank_cell(first_row["厚度"])
                self.assert_blank_cell(first_row["总数量"])

                self.assertEqual(second_parent_row["图号"], "100-B")
                self.assertEqual(second_parent_row["总数量"], "3")

                self.assertEqual(second_row["物料ID"], "CH-001")
                self.assertEqual(second_row["图号"], "PART-X")
                self.assertEqual(second_row["厚度"], "T2")
                self.assert_blank_cell(second_row["总数量"])
            finally:
                result_book.close()

    def test_summary_sheet_skips_note_rows_without_drawing_or_material_id(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            input_file = root / "input.xlsx"
            bom_dir = root / "boms"
            bom_dir.mkdir()
            output_file = root / "result.xlsx"

            create_workbook(
                input_file,
                [
                    ["序号", "图号", "名称", "数量"],
                    [8, "101002000250", "支架总成", 18],
                ],
            )

            create_workbook(
                bom_dir / "bom_note.xlsx",
                [
                    ["序号", "物料ID", "图号", "名称", "厚度", "材质", "数量", "总数量", "备注"],
                    [1, "", "101002000250（248010310）", "支架", "组件", "组件", 1, 1, "260118"],
                    ["1.1", "CH-001", "101002000251-A", "弯板", "T3", "06Cr19Ni10", 1, 1, ""],
                    ["", "", "", "工艺说明", "", "", "", "", "左右对称"],
                ],
                sheet_name="BOM",
            )

            run_search(input_file, bom_dir, output_file)

            result_book = load_workbook(output_file, data_only=True)
            try:
                aggregate_sheet = result_book["汇总BOM"]
                rows = list(aggregate_sheet.iter_rows(values_only=True))
                self.assertEqual(len(rows), 4)
                parent_row = self.row_map(rows[0], rows[1])
                material_row = self.row_map(rows[0], rows[2])
                note_row = self.row_map(rows[0], rows[3])
                self.assertEqual(parent_row["图号"], "101002000250(248010310)")
                self.assertEqual(material_row["物料ID"], "CH-001")
                self.assertEqual(material_row["图号"], "101002000251-A")
                self.assertEqual(material_row["名称"], "弯板")
                self.assertEqual(note_row["序号"], "8.2")
                self.assertEqual(note_row["名称"], "工艺说明")
            finally:
                result_book.close()


if __name__ == "__main__":
    unittest.main()
